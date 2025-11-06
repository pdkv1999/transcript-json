#!/usr/bin/env python3
"""
Entrypoint Flask app (thin) — imports functionality from modules in this package.
Run: python app.py
"""
import logging
import os
import time
import shutil
import json
from pathlib import Path
from flask import Flask, flash, render_template, redirect, request, url_for, send_from_directory
from markupsafe import Markup

from readers import read_transcript
from extractor import fill_schema_locally, fill_schema_with_gemini_then_local
from io_helpers import OUTPUT_DIR, UPLOAD_DIR, DATA_DIR, write_text_file

# NEW: analytics blueprint
from analytics_routes import analytics_bp

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET", "change-me")

logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s %(message)s"
)

# ---------------- Normalizers ---------------- #

def normalize_schema_to_json(src_path: Path, dst_json: Path) -> None:
    """
    Convert uploaded schema (JSON/CSV/TSV/TXT/XLS/XLSX) into JSON array of rows the LLM expects.
    Columns (case-insensitive, aliases allowed): row_key, row_label, tier6, full_row.
    """
    ext = src_path.suffix.lower()
    rows = []

    def _strip(v): return "" if v is None else str(v).strip()

    def coerce_row(d, i):
        rk = _strip(d.get("row_key") or d.get("Row Key") or d.get("key"))
        rl = _strip(d.get("row_label") or d.get("Row Label") or d.get("label") or d.get("name") or d.get("title"))
        t6 = _strip(d.get("tier6") or d.get("Tier6") or d.get("type") or d.get("field_type"))
        fr = d.get("full_row")
        if fr is None:
            fr = d.get("Full Row") or d.get("description") or d.get("details") or d.get("notes") or ""
        out = {"row_key": rk, "row_label": rl, "tier6": t6, "full_row": fr, "row_index": i}
        if not out["row_key"]:
            base = out["row_label"] or f"row_{i}"
            out["row_key"] = base.lower().replace(" ", "_")
        return out

    try:
        if ext == ".json":
            data = json.loads(src_path.read_text(encoding="utf-8"))
            if isinstance(data, dict) and "rows" in data: data = data["rows"]
            if isinstance(data, list):
                for i, r in enumerate(data): rows.append(coerce_row(dict(r or {}), i))

        elif ext in (".csv", ".tsv", ".txt"):
            import csv
            sample = src_path.read_text(encoding="utf-8", errors="replace")
            if ext == ".tsv":
                delimiter = "\t"
            elif ext == ".csv":
                delimiter = ","
            else:
                try:
                    dialect = csv.Sniffer().sniff(sample[:8192], delimiters=[",", "\t", ";", "|"])
                    delimiter = dialect.delimiter
                except Exception:
                    delimiter = ","
            try:
                has_header = csv.Sniffer().has_header(sample[:8192])
            except Exception:
                first = (sample.splitlines() or [""])[0].lower()
                has_header = any(h in first for h in ("row_key", "row label", "tier6", "full"))
            with src_path.open(newline="", encoding="utf-8", errors="replace") as f:
                if has_header:
                    reader = csv.DictReader(f, delimiter=delimiter)
                    for i, r in enumerate(reader): rows.append(coerce_row(dict(r or {}), i))
                else:
                    reader = csv.reader(f, delimiter=delimiter)
                    for i, cols in enumerate(reader):
                        cols = list(cols or [])
                        rec = {
                            "row_label": cols[0] if len(cols) > 0 else f"row_{i}",
                            "tier6": cols[1] if len(cols) > 1 else "",
                            "full_row": cols[2] if len(cols) > 2 else ""
                        }
                        rows.append(coerce_row(rec, i))

        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
            except Exception as e:
                raise RuntimeError(f"XLSX support requires openpyxl: {e}")
            wb = openpyxl.load_workbook(src_path, data_only=True)
            ws = wb.active
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [(_strip(c.value)).lower() if c.value is not None else "" for c in header_cells]
            start_row = 2 if any(headers) else 1
            if not any(headers): headers = ["row_label", "tier6", "full_row"]
            for i, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True)):
                rec = {}
                for j, h in enumerate(headers):
                    if not h: continue
                    rec[h] = row[j] if j < len(row) else ""
                rows.append(coerce_row(rec, i))
        else:
            rows = []

    except Exception as e:
        raise RuntimeError(f"Failed to parse schema {src_path.name}: {e}")

    dst_json.parent.mkdir(parents=True, exist_ok=True)
    dst_json.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_recs_to_json(src_path: Path, dst_json: Path) -> None:
    """
    Convert a Recommendations schema (XLSX/CSV/TSV/TXT/JSON) into a mapping:
    {
      "<Domain>": {
        "high": ["rec 1", "rec 2", ...],   # outcome Alot/Always
        "some": ["rec 3", ...]             # outcome Just Some
      },
      ...
    }
    Expected columns (case-insensitive, aliases allowed):
      - domain (aliases: area, category)
      - outcome (values like "A lot/Always", "Alot/Always", "Just Some"; we normalize to "high" or "some")
      - recommendation (aliases: rec, action, suggestion)
    """
    ext = src_path.suffix.lower()
    rules = {}

    def _strip(v): return "" if v is None else str(v).strip()

    def _lvl(val: str) -> str:
        v = (_strip(val)).lower()
        if "lot" in v or "always" in v or "a lot" in v or "alot" in v:
            return "high"
        if "some" in v or "just some" in v:
            return "some"
        # fallback: try to infer from keywords
        if "high" in v or "severe" in v: return "high"
        return "some"

    def add_rule(domain: str, lvl: str, rec: str):
        if not domain or not rec: return
        domain = domain.strip()
        lvl = "high" if lvl == "high" else "some"
        if domain not in rules: rules[domain] = {"high": [], "some": []}
        if rec not in rules[domain][lvl]:
            rules[domain][lvl].append(rec)

    try:
        if ext == ".json":
            data = json.loads(src_path.read_text(encoding="utf-8"))
            # Accept either mapping format already or list of dict rows
            if isinstance(data, dict):
                rules = data
            elif isinstance(data, list):
                for r in data:
                    d = _strip(r.get("domain") or r.get("area") or r.get("category"))
                    o = _lvl(r.get("outcome") or r.get("Outcome") or r.get("level"))
                    rec = _strip(r.get("recommendation") or r.get("rec") or r.get("action") or r.get("suggestion"))
                    add_rule(d, o, rec)

        elif ext in (".csv", ".tsv", ".txt"):
            import csv
            sample = src_path.read_text(encoding="utf-8", errors="replace")
            if ext == ".tsv":
                delimiter = "\t"
            elif ext == ".csv":
                delimiter = ","
            else:
                try:
                    dialect = csv.Sniffer().sniff(sample[:8192], delimiters=[",", "\t", ";", "|"])
                    delimiter = dialect.delimiter
                except Exception:
                    delimiter = ","
            with src_path.open(newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                for r in reader:
                    d = _strip(r.get("domain") or r.get("area") or r.get("category"))
                    o = _lvl(r.get("outcome") or r.get("Outcome") or r.get("level"))
                    rec = _strip(r.get("recommendation") or r.get("rec") or r.get("action") or r.get("suggestion"))
                    add_rule(d, o, rec)

        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
            except Exception as e:
                raise RuntimeError(f"XLSX support requires openpyxl: {e}")
            wb = openpyxl.load_workbook(src_path, data_only=True)
            ws = wb.active
            # Header row
            headers = [(_strip(c.value)).lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {h: i for i, h in enumerate(headers)}
            # Flexible header mapping
            def get(row, *names):
                for n in names:
                    j = idx.get(n.lower())
                    if j is not None and j < len(row):
                        return row[j]
                return None
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = _strip(get(row, "domain", "area", "category"))
                o = _lvl(get(row, "outcome", "level"))
                rec = _strip(get(row, "recommendation", "rec", "action", "suggestion"))
                add_rule(d, o, rec)

        else:
            rules = {}

    except Exception as e:
        raise RuntimeError(f"Failed to parse recommendations {src_path.name}: {e}")

    dst_json.parent.mkdir(parents=True, exist_ok=True)
    dst_json.write_text(json.dumps(rules, ensure_ascii=False, indent=2), encoding="utf-8")


# Register analytics blueprint
app.register_blueprint(analytics_bp)

@app.route("/", methods=["GET"])
def index():
    outputs = sorted([p.name for p in OUTPUT_DIR.glob("*_filled.*")], reverse=True)[:50]
    return render_template("index.html", outputs=outputs)

@app.route("/upload", methods=["POST"])
def upload():
    if "transcript" not in request.files or "schema" not in request.files:
        flash("Please upload both transcript and schema files.")
        return redirect(url_for("index"))

    tr_file = request.files["transcript"]
    sc_file = request.files["schema"]
    recs_file = request.files.get("recs")  # <-- NEW (optional)
    use_llm = bool(request.form.get("use_llm"))

    if not tr_file or tr_file.filename == "":
        flash("Transcript missing.")
        return redirect(url_for("index"))
    if not sc_file or sc_file.filename == "":
        flash("Schema missing.")
        return redirect(url_for("index"))

    # Per-run session id (used by analytics page)
    ts = int(time.time())
    session_id = str(ts)

    # Save uploads to /uploads for traceability
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    tr_fname = f"{ts}_{Path(tr_file.filename).name}"
    sc_fname = f"{ts}_{Path(sc_file.filename).name}"
    tr_path = UPLOAD_DIR / tr_fname
    sc_path = UPLOAD_DIR / sc_fname
    tr_file.save(tr_path)
    sc_file.save(sc_path)

    # Also persist copies into /data (non-fatal if it fails)
    try:
        shutil.copy2(tr_path, DATA_DIR / Path(tr_file.filename).name)
    except Exception:
        logging.exception("Failed to persist transcript copy to data/ (non-fatal).")
    try:
        shutil.copy2(sc_path, DATA_DIR / Path(sc_file.filename).name)
    except Exception:
        logging.exception("Failed to persist schema copy to data/ (non-fatal).")

    # Create session directory for analytics page
    session_dir = DATA_DIR / session_id
    try:
        session_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        logging.exception("Failed to create session directory: %s", session_dir)

    # Read transcript text once, save a session copy for analytics
    try:
        transcript_text = read_transcript(tr_path)
        (session_dir / "transcript.txt").write_text(transcript_text, encoding="utf-8")
    except Exception as e:
        logging.exception("Failed to read transcript: %s", e)
        flash(f"Failed to read transcript: {e}")
        return redirect(url_for("index"))

    # Normalize schema into JSON
    schema_rows_path = session_dir / "schema_rows.json"
    try:
        normalize_schema_to_json(sc_path, schema_rows_path)
    except Exception as e:
        logging.exception("Failed to normalize schema into JSON (non-fatal): %s", e)
        try:
            shutil.copy2(sc_path, session_dir / sc_path.name)
        except Exception:
            pass

    # OPTIONAL: normalize recommendations into JSON (if uploaded)
    if recs_file and recs_file.filename:
        recs_path = UPLOAD_DIR / f"{ts}_{Path(recs_file.filename).name}"
        recs_file.save(recs_path)
        try:
            normalize_recs_to_json(recs_path, session_dir / "recommendations.json")
        except Exception as e:
            logging.exception("Failed to normalize recommendations (non-fatal): %s", e)
            # keep original for debugging
            try:
                shutil.copy2(recs_path, session_dir / Path(recs_file.filename).name)
            except Exception:
                pass

    # Optional header meta for analytics
    try:
        meta_path = session_dir / "report_meta.json"
        if not meta_path.exists():
            meta_stub = {
                "child_name": request.form.get("child_name", "—"),
                "child_age": request.form.get("child_age", "—"),
                "date_of_interview": request.form.get("date_of_interview", time.strftime("%d %b %Y")),
                "parent_respondent": request.form.get("parent_respondent", "—"),
                "interviewer": request.form.get("interviewer", "—"),
                "referral_source": request.form.get("referral_source", "—"),
                "report_title": request.form.get("report_title", "Parent Telephone Interview Summary"),
            }
            meta_path.write_text(json.dumps(meta_stub, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        logging.exception("Failed to write report_meta.json (non-fatal).")

    # Run your existing pipeline (LLM or local)
    try:
        if use_llm:
            df, out_path = fill_schema_with_gemini_then_local(sc_path, transcript_text)
        else:
            df, out_path = fill_schema_locally(sc_path, transcript_text)
    except Exception as e:
        logging.exception("Processing error: %s", e)
        flash(f"Processing error: {e}")
        return redirect(url_for("index"))

    out_fname = out_path.name if out_path else "unknown"

    # Write debug files (best-effort)
    try:
        write_text_file("debug_transcript.txt", transcript_text[:300000])
    except Exception:
        pass

    # Link user directly to analytics page for this run
    analytics_url = url_for("analytics.page_analytics", session_id=session_id)
    flash(Markup(
        f"Processing complete. Saved: <code>data/{out_fname}</code> &nbsp; | &nbsp; "
        f"<a href='{analytics_url}' target='_blank' style='font-weight:bold;'>Open Analytics Report</a>"
    ))

    return render_template(
        "result.html",
        out_fname=out_fname,
        debug_transcript="debug_transcript.txt",
        debug_sentences="debug_sentences.txt",
        analytics_url=analytics_url
    )

@app.route("/data/<path:fname>")
def download(fname):
    p = OUTPUT_DIR / fname
    if not p.exists():
        return "Not found", 404
    return send_from_directory(str(OUTPUT_DIR), fname, as_attachment=True)

@app.route("/download/<path:fname>")
def download_alt(fname):
    return download(fname)

if __name__ == "__main__":
    logging.info("Starting modular app.")
    debug_mode = os.getenv("FLASK_DEBUG", "false").lower() in ("1", "true", "yes")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5050)), debug=debug_mode)
