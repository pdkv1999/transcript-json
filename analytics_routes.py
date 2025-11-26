"""
Analytics routes - updated:
- Produces canonical domain objects with both frequency-based and quote-based metrics.
- Computes frequency percent using weighted formula (always=2, some=1, never=0),
  normalized by the maximum possible weighted score (2 * total) so percent is in 0..100.
- Rounds percent, freq_share and quote_share to 1 decimal place to avoid long floats.
- Exposes both traffic lists and domain_scores for both modes.
- Adds a new POST /api/analytics/upload endpoint that accepts a file upload
  (CSV/TSV/JSON/XLSX) and returns the same JSON analytics payload.
- Cleans up uploaded temporary files after processing (best-effort).
- Includes deduplicated unique quotes and concatenated all_quotes_text per domain.
- Stricter dedupe (threshold 0.90) and caps returned unique quotes to 12 per domain.
- NEW: Looks up recommendations for each domain based on
  Parent-Friendly Domain + Concern Level from recommendations Excel/CSV
  and injects them into rows["recommendations"] for the UI.
- NEW: Adds functional impact fields for each item/row:
    * disrupts_eating_feeding
    * disrupts_social_activity
    * disrupts_major_role_obligations
    * disrupts_education
    * disrupts_social_obligations
    * disrupts_sleep
    * disrupts_physical_activity
    * disrupts_other_everyday_functioning
  and LLM helpers to generate y/n values when producing Excel/exports.
- NEW: Summary template matches Example 1 using:
    [Strengths], [Concern areas], [Class/year], [Domains of no concerns]
"""

import csv
import json
import re
import uuid
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, Tuple, List

from flask import Blueprint, jsonify, render_template, current_app, request

import difflib
from google.generativeai import GenerativeModel
import google.generativeai as genai
import os


analytics_bp = Blueprint("analytics", __name__, template_folder="templates")

DATA_ROOT = Path("data")
DATA_ROOT.mkdir(parents=True, exist_ok=True)

# -------------------------------------------------------------------
# NEW: Recommendations lookup configuration and helpers
# -------------------------------------------------------------------
RECOMMENDATIONS_ROOT = Path("recommendations")
RECOMMENDATIONS_ROOT.mkdir(parents=True, exist_ok=True)

_RECS_CACHE: Dict[Tuple[str, str], str] = {}
_RECS_CACHE_LOADED: bool = False

def _get_gemini_model_name() -> Optional[str]:
    """
    Use google.generativeai.list_models() to find a valid text model that supports
    generateContent. Prefer Gemini text models (pro/flash), fall back to the first
    text-capable model. Returns the full model name (e.g. 'models/gemini-1.5-pro-latest').
    """
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        current_app.logger.warning("Gemini: GEMINI_API_KEY not set; cannot choose model.")
        return None

    try:
        genai.configure(api_key=api_key)
        models = list(genai.list_models())
    except Exception as e:
        current_app.logger.exception("Gemini: list_models failed: %s", e)
        return None

    if not models:
        current_app.logger.warning("Gemini: list_models returned no models.")
        return None

    # Prefer models that support generateContent
    def supports_generate_content(m) -> bool:
        methods = getattr(m, "supported_generation_methods", []) or []
        return any("generatecontent" in x.lower() for x in methods)

    text_models = [m for m in models if supports_generate_content(m)]
    if not text_models:
        text_models = models  # fallback: at least pick something

    # Prefer non-embedding Gemini models with "pro" or "flash" in the name
    preferred = None
    for m in text_models:
        name = (getattr(m, "name", "") or "").lower()
        if "gemini" in name and "embed" not in name:
            if "pro" in name or "flash" in name:
                preferred = m
                break

    if not preferred and text_models:
        preferred = text_models[0]

    if not preferred:
        current_app.logger.warning("Gemini: no suitable model found in list_models.")
        return None

    model_name = getattr(preferred, "name", None)
    current_app.logger.info("Gemini: using model %s", model_name)
    return model_name

def _norm_domain_for_lookup(name: str) -> str:
    return (name or "").strip().lower()


def _norm_concern_for_lookup(label: str) -> str:
    """
    Normalise concern labels from the Excel (e.g. 'High concern', 'Some concern', 'No concern')
    into internal codes: 'hi' | 'some' | 'no'
    """
    s = (label or "").strip().lower()
    if "high" in s:
        return "hi"
    if "some" in s or "medium" in s:
        return "some"
    if "no" in s or "low" in s or "none" in s:
        return "no"
    # already short codes?
    if s in ("hi", "high"):
        return "hi"
    if s in ("some", "medium"):
        return "some"
    if s in ("no", "low", "none"):
        return "no"
    return s


# canonical keys (keep domain_* keys aligned with DOMAIN_CANON below)
CANONICAL_KEYS = [
    "summary_overview",
    "parent_quote",
    "traffic_high",
    "traffic_some",
    "traffic_no",
    "recommendations",
    "domain_anxiety_emotion",
    "domain_attention_adhd",
    "domain_autism",
    "domain_behaviour",
    "domain_care_history",
    "domain_developmental_history",
    "domain_eating",
    "domain_home_life_routine",
    "domain_learning_dyslexia",
    "domain_motor_skills",
    "domain_other",
    "domain_parents_goal",
    "domain_school_life",
    "domain_sleep",
    "domain_social_communication",
    "domain_strengths_interests",
    # NEW: functional impact flags (per item/row)
    "disrupts_eating_feeding",
    "disrupts_social_activity",
    "disrupts_major_role_obligations",
    "disrupts_education",
    "disrupts_social_obligations",
    "disrupts_sleep",
    "disrupts_physical_activity",
    "disrupts_other_everyday_functioning",
]

# authoritative list + display names / order
DOMAIN_CANON = [
    "Anxiety / Emotion",
    "Attention / ADHD",
    "Autism",
    "Behaviour",
    "Care History",
    "Developmental History",
    "Eating",
    "Home Life / Routine",
    "Learning / Dyslexia",
    "Motor Skills",
    "Other",
    "Parent's Goal for this Assessment",
    "School Life",
    "Sleep",
    "Social / Communication",
    "Strengths / Interests",
]

# keywords used to auto-assign raw rows into domains — add or tune as needed
DOMAIN_KEYWORDS = {
    "Anxiety / Emotion": [
        "anxi",
        "worry",
        "emotion",
        "mood",
        "sad",
        "tear",
        "fear",
        "panic",
        "stress",
    ],
    "Attention / ADHD": [
        "attention",
        "adhd",
        "hyper",
        "focus",
        "distract",
        "concentrat",
    ],
    "Autism": [
        "autis",
        "asperg",
        "stereotyp",
        "sensory",
        "repetit",
        "meltdo",
        "social communication",
    ],
    "Behaviour": [
        "behaviour",
        "behavior",
        "behav",
        "tantrum",
        "aggress",
        "rule",
        "routine",
        "challeng",
    ],
    "Care History": [
        "care history",
        "caregiver",
        "foster",
        "placement",
        "care history",
        "guardian",
    ],
    "Developmental History": [
        "development",
        "milestone",
        "delay",
        "developmental",
        "walk",
        "talk",
        "sit",
        "crawl",
    ],
    "Eating": [
        "eat",
        "feeding",
        "food",
        "weight",
        "appetite",
        "swallow",
        "mealtime",
    ],
    "Home Life / Routine": [
        "home",
        "routine",
        "bedtime routine",
        "family",
        "siblings",
        "house",
        "routine",
    ],
    "Learning / Dyslexia": [
        "learn",
        "dyslex",
        "reading",
        "spelling",
        "school",
        "education",
        "homework",
    ],
    "Motor Skills": [
        "motor",
        "coordina",
        "climb",
        "fine motor",
        "gross motor",
        "hand",
        "balance",
    ],
    "Other": [
        "other",
        "misc",
        "additional",
        "note",
        "info",
    ],
    "Parent's Goal for this Assessment": [
        "goal",
        "aim",
        "parents goal",
        "parent's goal",
        "objective",
        "what parent wants",
    ],
    "School Life": [
        "school",
        "class",
        "teacher",
        "peer",
        "homework",
        "attendance",
    ],
    "Sleep": [
        "sleep",
        "bed",
        "night",
        "insomnia",
        "nap",
        "tired",
    ],
    "Social / Communication": [
        "social",
        "friend",
        "communic",
        "talk",
        "speech",
        "language",
        "play",
        "peer",
    ],
    "Strengths / Interests": [
        "strength",
        "interest",
        "likes",
        "hobby",
        "talent",
        "strengths",
    ],
}


def _canonical_domain_from_label(label: str) -> Optional[str]:
    """
    Map a free-text domain label from the Excel (e.g. 'Anxiety/emotions',
    'School', 'Social & communication') to the closest canonical domain name
    in DOMAIN_CANON.
    """
    if not label:
        return None

    raw = str(label).strip()
    s = raw.lower()

    # Exact / substring matches first
    for canon in DOMAIN_CANON:
        c_low = canon.lower()
        if s == c_low:
            return canon
        if s in c_low or c_low in s:
            return canon

    # Very small alias normalisations
    alias_map = {
        "anxiety": "Anxiety / Emotion",
        "anxiety / emotions": "Anxiety / Emotion",
        "anxiety/emotion": "Anxiety / Emotion",
        "attention": "Attention / ADHD",
        "adhd": "Attention / ADHD",
        "school": "School Life",
        "school life / learning": "School Life",
        "learning": "Learning / Dyslexia",
        "reading / dyslexia": "Learning / Dyslexia",
        "social": "Social / Communication",
        "social communication": "Social / Communication",
        "motor": "Motor Skills",
        "motor skills / coordination": "Motor Skills",
        "sleep & routine": "Sleep",
        "home life": "Home Life / Routine",
        "strengths": "Strengths / Interests",
    }
    if s in alias_map:
        return alias_map[s]

    # Fallback: closest fuzzy match
    best = None
    best_score = 0.0
    for canon in DOMAIN_CANON:
        score = difflib.SequenceMatcher(None, s, canon.lower()).ratio()
        if score > best_score:
            best_score = score
            best = canon

    if best and best_score >= 0.55:
        return best

    return None


# ---------------------- helpers ----------------------


def _flatten_cell_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).strip()


def _empty_row() -> Dict[str, Any]:
    return {"value": None, "quote": None, "confidence": 0.0}


# ----------------------- File discovery and readers -----------------------


def _find_filled_file_for_session(session_id: str) -> Optional[Path]:
    session_dir = DATA_ROOT / session_id
    candidates: List[Path] = []
    if session_dir.exists() and session_dir.is_dir():
        for p in session_dir.iterdir():
            if p.is_file() and "_filled" in p.name:
                candidates.append(p)
    for p in DATA_ROOT.glob(f"*{session_id}*_filled.*"):
        if p.is_file():
            candidates.append(p)
    if not candidates:
        for p in DATA_ROOT.glob("*_filled.*"):
            if p.is_file():
                candidates.append(p)
    if not candidates:
        return None
    order = [".csv", ".xlsx", ".json", ".xls"]
    for ext in order:
        for c in candidates:
            if c.suffix.lower() == ext:
                return c
    return candidates[0]


def _read_csv_filled(path: Path) -> List[Dict[str, Any]]:
    text = path.read_text(encoding="utf-8", errors="replace")
    sample = text[:8192]
    delimiter = ","
    if "\t" in sample and sample.count("\t") > sample.count(","):
        delimiter = "\t"
    import csv as _csv
    f = path.open(newline="", encoding="utf-8", errors="replace")
    try:
        sniffer = _csv.Sniffer()
        try:
            dialect = sniffer.sniff(sample, delimiters=[",", "\t", ";", "|"])
            delimiter = dialect.delimiter
        except Exception:
            pass
    finally:
        f.close()
    rows = []
    with path.open(newline="", encoding="utf-8", errors="replace") as fh:
        reader = _csv.DictReader(fh, delimiter=delimiter)
        for r in reader:
            low = {(k or "").strip().lower(): (v if v is not None else "") for k, v in r.items()}
            rows.append(low)
    return rows


def _read_json_filled(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        out = []
        is_map = all(isinstance(v, (dict, str, int, float, type(None))) for v in data.values())
        if is_map and not isinstance(next(iter(data.values())), list):
            for rk, v in data.items():
                rec = {"row_key": rk}
                if isinstance(v, dict):
                    rec["value"] = v.get("value")
                    rec["quote"] = v.get("quote")
                    rec["confidence"] = v.get("confidence")
                else:
                    rec["value"] = v
                out.append(rec)
            return out
        if "rows" in data and isinstance(data["rows"], list):
            return data["rows"]
    if isinstance(data, list):
        return data
    return []


def _read_xlsx_filled(path: Path) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except Exception as e:
        current_app.logger.error("openpyxl required to read xlsx: %s", e)
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [(str(h).strip().lower() if h is not None else "") for h in header_row]
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            rec[h] = row[i] if i < len(row) else None
        rows.append(rec)
    return rows


# ----------------------- Mapping helpers -----------------------


def _canonical_row_key_from_label(label: str) -> Optional[str]:
    """
    Map human-readable question text from the Excel to canonical keys
    we can later use for the header + summary + functional impact.
    """
    if not label:
        return None

    s = str(label).strip().lower()
    s = s.replace("’", "'")  # normalise apostrophes

    # Basic header fields
    if "what is your child's full name" in s or "what is your childs full name" in s:
        return "child_name"

    if "how old is your child" in s:
        return "child_age"

    if "what is your full name" in s:
        # parent respondent name
        return "parent_respondent"

    if "who is conducting today's interview" in s or "who is conducting today" in s:
        return "interviewer"

    if "how were you referred to our service" in s or "how were you referred to the service" in s:
        return "referral_source"

    # NEW: class / year at school
    if "class/year" in s or "class / year" in s or "what class" in s or "what year" in s:
        return "child_class_year"

    # Domains we want for summary
    if "what are your main concerns about your child" in s or "main concerns about your child" in s:
        return "domain_parents_goal"

    if "what positive activities describe your child" in s or "positive activities describe your child" in s:
        return "domain_strengths_interests"

    if "has your child undergone any previous assessments" in s or "previous assessments" in s:
        return "domain_care_history"

    # NEW: functional impact fields (per item/row)
    if "disrupts eating" in s or "disrupts feeding" in s:
        return "disrupts_eating_feeding"

    if "disrupts social activity" in s:
        return "disrupts_social_activity"

    if "disrupts fulfillment of major role obligations" in s or "major role obligations" in s:
        return "disrupts_major_role_obligations"

    if "disrupts education" in s:
        return "disrupts_education"

    if "disrupts fulfillment of social obligations" in s or "social obligations" in s:
        return "disrupts_social_obligations"

    if "disrupts sleep" in s:
        return "disrupts_sleep"

    if "disrupts physical activity" in s:
        return "disrupts_physical_activity"

    if "disrupts other everyday functioning" in s or "other everyday functioning" in s:
        return "disrupts_other_everyday_functioning"

    return None


def _map_row_list_to_dict(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for r in rows:

        def get_any(*names):
            for n in names:
                if n in r and r[n] not in (None, ""):
                    return r[n]
            return None

        row_key = get_any("row_key", "key", "row", "rowlabel", "row_label", "row label")
        row_label = get_any(
            "row_label",
            "label",
            "rowlabel",
            "row label",
            "possible parent-friendly question",
            "question",
        )
        value = get_any(
            "value",
            "answer",
            "filled_value",
            "response",
            "result",
            "frequency",
            "frequency_code",
            "frequency code",
            "response_code",
        )
        quote = get_any("quote", "evidence", "supporting_quote", "support", "found_quote")
        conf = get_any("confidence", "conf", "score", "certainty")

        # 🔹 NEW: infer canonical row_key from the human-readable question text
        if row_label:
            canon = _canonical_row_key_from_label(str(row_label))
            if canon:
                row_key = canon

        if not row_key and row_label:
            row_key = str(row_label).strip().lower().replace(" ", "_")

        if isinstance(conf, str):
            try:
                conf_num = float(conf.strip())
            except Exception:
                conf_num = 0.0
        elif isinstance(conf, (int, float)):
            conf_num = float(conf)
        else:
            conf_num = 0.0

        key = (
            str(row_key)
            if row_key
            else (str(row_label).strip().lower().replace(" ", "_") if row_label else None)
        )
        if not key:
            for kk, vv in r.items():
                if vv not in (None, ""):
                    key = kk.strip().lower() + "_" + str(abs(hash(str(vv))))[:6]
                    break
        if not key:
            continue

        val_s = None if value is None or (isinstance(value, str) and value.strip() == "") else value
        quote_s = None if quote is None or (isinstance(quote, str) and quote.strip() == "") else quote
        out[key] = {"value": val_s, "quote": quote_s, "confidence": conf_num}
    return out


# ----------------------- Frequency normalization and scoring -----------------------


def _normalize_freq_token(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    s = _flatten_cell_value(v).lower()
    s = s.strip()
    if s == "":
        return None
    if s in ("2", "2.0", "high"):
        return "always"
    if s in ("1", "1.0", "some"):
        return "some"
    if s in ("0", "0.0", "none", "no", "never"):
        return "never"
    if any(x in s for x in ("a lot", "alot", "always", "frequent", "often", "many")):
        return "always"
    if any(x in s for x in ("some", "just some", "occasionally", "sometimes")):
        return "some"
    if any(x in s for x in ("notmuch", "not much", "never", "no", "none")):
        return "never"
    m = re.search(r"\b([0-9])\b", s)
    if m:
        if m.group(1) == "2":
            return "always"
        if m.group(1) == "1":
            return "some"
        if m.group(1) == "0":
            return "never"
    return None


def _token_weight(tok: Optional[str]) -> int:
    if tok == "always":
        return 2
    if tok == "some":
        return 1
    return 0


# ----------------------- Domain helpers -----------------------


def _domain_label_from_row_key_or_text(text: str) -> str:
    s = _flatten_cell_value(text).lower()
    for canon, kws in DOMAIN_KEYWORDS.items():
        for kw in kws:
            if kw in s:
                return canon
    for canon in DOMAIN_CANON:
        if any(tok in canon.lower() for tok in s.split()):
            return canon
    return text.strip().title() if text else "Unknown"


def _find_domain_rows_in_raw_rows(rows_list: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Group raw rows by domain.

    Priority:
      1. If the row has an explicit domain column
         (e.g. 'Parent-Friendly Domain', 'Domain'), use that.
      2. Otherwise, fall back to keyword / label-based guessing.
    """
    domain_candidates: Dict[str, List[Dict[str, Any]]] = {d: [] for d in DOMAIN_CANON}

    for r in rows_list:
        # ---------- 1) Try explicit domain column ----------
        explicit_dom = None
        for col in (
            "Parent-Friendly Domain",
            "parent-friendly domain",
            "Parent friendly domain",
            "parent friendly domain",
            "Domain",
            "domain",
        ):
            if col in r and r[col] not in (None, ""):
                explicit_dom = _canonical_domain_from_label(r[col])
                break

        if explicit_dom:
            domain_candidates[explicit_dom].append(r)
            continue  # do not re-guess – we trust the explicit domain

        # ---------- 2) Fallback: keyword / label detection ----------
        combined_text = " ".join([_flatten_cell_value(r.get(k)) for k in r.keys() if r.get(k)])
        combined_text_lower = combined_text.lower()
        assigned = set()

        # a) keyword scan
        for domain, kws in DOMAIN_KEYWORDS.items():
            for kw in kws:
                if kw in combined_text_lower:
                    domain_candidates[domain].append(r)
                    assigned.add(domain)
                    break

        if assigned:
            continue

        # b) look at row label / question text
        for k in ("row_key", "row_label", "label", "question", "possible parent-friendly question"):
            if k in r and r[k]:
                dom = _canonical_domain_from_label(str(r[k]))
                if dom and dom in domain_candidates:
                    domain_candidates[dom].append(r)
                    assigned.add(dom)
                    break

        # If still nothing, leave unassigned (row won't influence any domain)
    return domain_candidates


# ----------------------- Core loader -----------------------


def load_filled_rows_for_session(
    session_id: str,
) -> Tuple[Dict[str, Dict[str, Any]], Optional[Path], List[Dict[str, Any]]]:
    p = _find_filled_file_for_session(session_id)
    if not p:
        return {}, None, []
    current_app.logger.info("Analytics: reading filled file: %s", p)
    rows_list: List[Dict[str, Any]] = []
    try:
        if p.suffix.lower() in (".csv", ".txt", ".tsv"):
            rows_list = _read_csv_filled(p)
        elif p.suffix.lower() in (".json",):
            rows_list = _read_json_filled(p)
        elif p.suffix.lower() in (".xlsx", ".xls"):
            rows_list = _read_xlsx_filled(p)
        else:
            rows_list = _read_csv_filled(p)
    except Exception as e:
        current_app.logger.exception("Analytics: failed to read filled file %s: %s", p, e)
        return {}, p, []
    mapped = _map_row_list_to_dict(rows_list)

    normalized: Dict[str, Dict[str, Any]] = {}
    for k in CANONICAL_KEYS:
        if k in mapped:
            normalized[k] = mapped[k]
            continue
        found = None
        for mk in mapped.keys():
            if k in mk:
                found = mk
                break
        if found:
            normalized[k] = mapped[found]
            continue
        words = k.replace("_", " ").split()
        best = None
        for mk in mapped.keys():
            if all(w.lower() in mk.lower() for w in words[:2]):
                best = mk
                break
        if best:
            normalized[k] = mapped[best]
            continue
        normalized[k] = _empty_row()

    # domain matching and population of canonical domain_* keys
    domain_candidates = _find_domain_rows_in_raw_rows(rows_list)

    domain_key_map = {
        "domain_anxiety_emotion": "Anxiety / Emotion",
        "domain_attention_adhd": "Attention / ADHD",
        "domain_autism": "Autism",
        "domain_behaviour": "Behaviour",
        "domain_care_history": "Care History",
        "domain_developmental_history": "Developmental History",
        "domain_eating": "Eating",
        "domain_home_life_routine": "Home Life / Routine",
        "domain_learning_dyslexia": "Learning / Dyslexia",
        "domain_motor_skills": "Motor Skills",
        "domain_other": "Other",
        "domain_parents_goal": "Parent's Goal for this Assessment",
        "domain_school_life": "School Life",
        "domain_sleep": "Sleep",
        "domain_social_communication": "Social / Communication",
        "domain_strengths_interests": "Strengths / Interests",
    }

    # helper pick best candidate (prefers explicit freq/quote)
    def _pick_best_candidate(cands: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        if not cands:
            return None
        best = None
        best_score = -1
        for r in cands:
            score = 0
            for fk in (
                "frequency_code",
                "frequency code",
                "frequency",
                "freq",
                "response_code",
                "response code",
                "value",
                "answer",
            ):
                v = r.get(fk) if fk in r else None
                if v not in (None, "", []):
                    score += 5
                    break
            for qk in ("quote", "evidence", "support", "found_quote", "supporting_quote"):
                if qk in r and r[qk] not in (None, ""):
                    score += 3
            if any(k in r and r[k] not in (None, "") for k in ("row_label", "label", "question")):
                score += 1
            combined = " ".join([_flatten_cell_value(r.get(k)) for k in r.keys() if r.get(k)])
            if 10 < len(combined) < 800:
                score += 1
            if score > best_score:
                best_score = score
                best = r
        return best

    for dk, domain_name in domain_key_map.items():
        existing = normalized.get(dk) or {}
        if existing.get("value") not in (None, "") or existing.get("quote") not in (None, ""):
            continue
        cands = domain_candidates.get(domain_name, []) or []
        chosen = _pick_best_candidate(cands)
        if chosen:
            val = None
            quote = None
            for fk in (
                "frequency_code",
                "frequency code",
                "frequency",
                "freq",
                "response_code",
                "response code",
                "value",
                "answer",
            ):
                if fk in chosen and chosen[fk] not in (None, ""):
                    val = chosen[fk]
                    break
            for qk in ("quote", "evidence", "support", "found_quote", "supporting_quote"):
                if qk in chosen and chosen[qk] not in (None, ""):
                    quote = chosen[qk]
                    break
            if val is None:
                for hk in ("row_label", "label", "question", "possible parent-friendly question"):
                    if hk in chosen and chosen[hk] not in (None, ""):
                        val = chosen[hk]
                        break
            if quote is None:
                long_text = None
                for v in chosen.values():
                    if v and isinstance(v, str) and len(v) > 30 and len(v) < 1000:
                        long_text = v
                        break
                if long_text:
                    quote = long_text
            normalized[dk] = {
                "value": val if val not in (None, "") else None,
                "quote": quote if quote not in (None, "") else None,
                "confidence": 1.0 if val or quote else 0.0,
            }

    return normalized, p, rows_list


# ----------------------- Unique-quote dedupe -----------------------


def _normalize_quote_text(q: str) -> str:
    s = q or ""
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r'[“”"\'\u2018\u2019]', "", s)
    s = s.lower().strip()
    return s


# NEW: helpers to detect / strip "inferred" text
def _is_inferred_text(text: Optional[str]) -> bool:
    if not text:
        return False
    return bool(re.search(r"\binferred\b", str(text), flags=re.IGNORECASE))


def _strip_inferred_text(text: Optional[str]) -> Optional[str]:
    """
    For summary fields: remove any 'inferred ...' part.
    If the whole thing is inferred, return None.
    """
    if not text:
        return None
    s = str(text)
    if not _is_inferred_text(s):
        s = s.strip()
        return s or None

    # If "inferred" appears, drop everything from that word onwards
    parts = re.split(r"\binferred\b", s, flags=re.IGNORECASE)
    cleaned = parts[0].strip()
    return cleaned or None


def _dedupe_quotes(quotes: List[str], threshold: float = 0.90) -> List[str]:
    """Dedupe by similarity >= threshold using difflib.SequenceMatcher."""
    unique: List[str] = []
    for q in quotes:
        nq = _normalize_quote_text(q)
        if not nq:
            continue
        is_dup = False
        for u in unique:
            sim = difflib.SequenceMatcher(None, nq, _normalize_quote_text(u)).ratio()
            if sim >= threshold:
                is_dup = True
                break
        if not is_dup:
            unique.append(q)
    return unique


# ----------------------- Aggregation / scoring -----------------------


def _aggregate_domain_weights_from_rows(rows_list: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Aggregate frequency + quote stats per domain.

    - Domain assignment:
        1) use explicit Domain / Parent-Friendly Domain column if present
        2) otherwise fall back to keyword / label guessing
    - Quotes are split into:
        * YES quotes   -> used in charts / percentages.
        * NO  quotes   -> shown in UI but NOT used in maths.
    - Any text containing 'inferred' is ignored completely.
    """
    domain_buckets = {
        d: {
            "always": 0,
            "some": 0,
            "never": 0,
            "unknown": 0,
            "yes_quotes_raw": [],
            "no_quotes_raw": [],
        }
        for d in DOMAIN_CANON
    }

    # we still pre-compute candidates for the guessing fallback
    domain_candidates = _find_domain_rows_in_raw_rows(rows_list)

    for r in rows_list:
        # ---------- 1) try explicit domain column ----------
        domain = None
        for col in (
            "Parent-Friendly Domain",
            "parent-friendly domain",
            "Parent friendly domain",
            "parent friendly domain",
            "Domain",
            "domain",
        ):
            if col in r and r[col] not in (None, ""):
                domain = _canonical_domain_from_label(r[col])
                break

        # ---------- 2) fallback: reuse candidate mapping / keyword logic ----------
        if not domain:
            combined_text = " ".join([_flatten_cell_value(r.get(k)) for k in r.keys() if r.get(k)])
            txt_lower = combined_text.lower()

            # a) direct keyword scan
            for dom, kws in DOMAIN_KEYWORDS.items():
                for kw in kws:
                    if kw in txt_lower:
                        domain = dom
                        break
                if domain:
                    break

        if not domain:
            # b) see if this row is one of the pre-grouped candidates
            for dom, cand_list in domain_candidates.items():
                if r in cand_list:
                    domain = dom
                    break

        if not domain:
            # c) final best-hit keyword fallback
            best_dom = None
            best_hits = 0
            combined_text = " ".join([_flatten_cell_value(r.get(k)) for k in r.keys() if r.get(k)])
            txt_lower = combined_text.lower()
            for dom, kws in DOMAIN_KEYWORDS.items():
                hits = sum(1 for kw in kws if kw in txt_lower)
                if hits > best_hits:
                    best_hits = hits
                    best_dom = dom
            if best_dom and best_hits > 0:
                domain = best_dom

        if not domain:
            # cannot assign this row to any domain – skip it
            continue

        # ---------- Frequency aggregation ----------
        combined_text = " ".join([_flatten_cell_value(r.get(k)) for k in r.keys() if r.get(k)])
        freq_raw = None
        for fk in (
            "frequency_code",
            "frequency code",
            "frequency",
            "freq",
            "response_code",
            "response code",
            "value",
            "answer",
        ):
            if fk in r and r[fk] not in (None, ""):
                freq_raw = r[fk]
                break
        if freq_raw is None:
            freq_raw = combined_text

        norm = _normalize_freq_token(freq_raw)
        if norm is None:
            domain_buckets[domain]["unknown"] += 1
        else:
            domain_buckets[domain].setdefault(norm, 0)
            domain_buckets[domain][norm] += 1

        # ---------- YES/NO flag for quotes ----------
        quote_flag = None  # 'yes', 'no', or None

        # 1) explicit "use quote" style columns
        for fk in (
            "use_quotes",
            "use quotes",
            "use_quote",
            "use quote",
            "include_quote",
            "include quote",
            "use_in_report",
            "use in report",
            "use this quote",
            "use this quote?",
        ):
            if fk in r and r[fk] not in (None, ""):
                val = _flatten_cell_value(r[fk]).lower()
                if val.startswith("y") or val == "1" or val == "yes":
                    quote_flag = "yes"
                elif val.startswith("n") or val == "0" or val == "no":
                    quote_flag = "no"
                break

        # 2) fallback to response/response_code if they look like Yes/No
        if quote_flag is None:
            for fk in ("response_code", "response", "quote_response", "quote_response_code"):
                if fk in r and r[fk] not in (None, ""):
                    val = _flatten_cell_value(r[fk]).lower()
                    if val in ("yes", "y", "no", "n"):
                        quote_flag = "yes" if val in ("yes", "y") else "no"
                        break

        # default: if no explicit flag, treat as YES so you don't lose older data
        if quote_flag is None:
            quote_flag = "yes"

        # ---------- collect quotes (no 'inferred') ----------
        for qk in ("quote", "evidence", "support", "found_quote", "supporting_quote"):
            if qk in r and r[qk] not in (None, ""):
                text = str(r[qk]).strip()
                if not text:
                    continue
                if _is_inferred_text(text):
                    # completely drop 'inferred' statements
                    continue
                if quote_flag == "no":
                    domain_buckets[domain]["no_quotes_raw"].append(text)
                else:
                    domain_buckets[domain]["yes_quotes_raw"].append(text)

    # ---------- build aggregated results ----------
    results: Dict[str, Dict[str, Any]] = {}
    weighted_totals_sum = 0

    for domain, counts in domain_buckets.items():
        a = counts.get("always", 0)
        s = counts.get("some", 0)
        n = counts.get("never", 0)
        total = a + s + n
        weighted_sum = 2 * a + 1 * s

        if total > 0:
            percent = (weighted_sum / float(2 * total)) * 100.0
        else:
            percent = 0.0
        percent = round(percent, 1)

        yes_unique = _dedupe_quotes(counts.get("yes_quotes_raw", []))
        no_unique = _dedupe_quotes(counts.get("no_quotes_raw", []))

        MAX_QUOTES_PER_DOMAIN = 12
        yes_trunc = yes_unique[:MAX_QUOTES_PER_DOMAIN]
        no_trunc = no_unique[:MAX_QUOTES_PER_DOMAIN]

        q_yes_count = len(yes_unique)
        q_no_count = len(no_unique)

        results[domain] = {
            "counts": {
                "always": a,
                "some": s,
                "never": n,
                "unknown": counts.get("unknown", 0),
            },
            "weighted_sum": weighted_sum,
            "total_count": total,
            "percent": percent,
            "yes_quotes": yes_trunc,
            "no_quotes": no_trunc,
            "quote_count": q_yes_count,          # YES only – used for maths
            "quote_yes_count": q_yes_count,
            "quote_no_count": q_no_count,
            "quote_total_text": " ".join(yes_unique),
            # note: quote_share & freq_share filled below
        }

        weighted_totals_sum += weighted_sum

    # frequency share
    if weighted_totals_sum > 0:
        for domain, info in results.items():
            info["freq_share"] = round(
                (float(info["weighted_sum"]) / float(weighted_totals_sum)) * 100.0, 1
            )
    else:
        for domain in results:
            results[domain]["freq_share"] = 0.0

    # quote share – YES quotes only
    total_quote_count = sum(info["quote_count"] for info in results.values())
    if total_quote_count > 0:
        for domain, info in results.items():
            info["quote_share"] = round(
                (float(info["quote_count"]) / float(total_quote_count)) * 100.0, 1
            )
    else:
        for domain in results:
            results[domain]["quote_share"] = 0.0

    return results


def _percent_to_level(percent: float) -> str:
    if percent is None:
        return "no"
    try:
        p = float(percent)
    except Exception:
        return "no"
    if p > 70:
        return "hi"
    if p >= 30:
        return "some"
    if p < 30:
        return "no"
    return "no"


# ----------------------- Recommendations loader -----------------------


def _load_recommendations_table() -> Dict[Tuple[str, str], str]:
    """
    Load mapping: (Parent-Friendly Domain, Concern Level) -> Recommendations text.
    Uses the first XLSX/CSV it finds in DATA_ROOT/recommendations.
    """
    global _RECS_CACHE_LOADED, _RECS_CACHE
    if _RECS_CACHE_LOADED:
        return _RECS_CACHE

    _RECS_CACHE_LOADED = True
    _RECS_CACHE = {}

    try:
        if not RECOMMENDATIONS_ROOT.exists():
            current_app.logger.warning("Recommendations root %s does not exist", RECOMMENDATIONS_ROOT)
            return _RECS_CACHE

        candidates: List[Path] = []
        for p in RECOMMENDATIONS_ROOT.iterdir():
            if p.suffix.lower() in (".xlsx", ".xls", ".csv", ".tsv", ".txt"):
                candidates.append(p)

        if not candidates:
            current_app.logger.warning("No recommendations file found under %s", RECOMMENDATIONS_ROOT)
            return _RECS_CACHE

        # Prefer xlsx
        order = {".xlsx": 0, ".xls": 1, ".csv": 2, ".tsv": 3, ".txt": 4}
        candidates.sort(key=lambda p: order.get(p.suffix.lower(), 99))
        path = candidates[0]
        current_app.logger.info("Using recommendations file: %s", path)

        if path.suffix.lower() in (".xlsx", ".xls"):
            try:
                import openpyxl
            except Exception as e:
                current_app.logger.error("openpyxl required for recommendations: %s", e)
                return _RECS_CACHE

            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header_map = {
                (str(h).strip().lower() if h is not None else ""): idx
                for idx, h in enumerate(header_row)
            }

            def find_col(*cands):
                for key, idx in header_map.items():
                    for cand in cands:
                        if cand in key:
                            return idx
                return None

            dom_idx = find_col("parent-friendly domain", "parent friendly", "domain")
            level_idx = find_col("concern level", "concern")
            rec_idx = find_col(
                "recommendations, supports, and resources",
                "recommendations",
                "resources",
                "supports",
            )

            if dom_idx is None or level_idx is None or rec_idx is None:
                current_app.logger.warning(
                    "Recommendations file missing expected columns (domain/concern/recommendations)"
                )
                return _RECS_CACHE

            for row in ws.iter_rows(min_row=2, values_only=True):
                dom = row[dom_idx] if dom_idx < len(row) else None
                lvl = row[level_idx] if level_idx < len(row) else None
                rec = row[rec_idx] if rec_idx < len(row) else None
                if not dom or not lvl or not rec:
                    continue
                dom_key = _norm_domain_for_lookup(str(dom))
                lvl_key = _norm_concern_for_lookup(str(lvl))
                text = str(rec).strip()
                if not text:
                    continue
                _RECS_CACHE[(dom_key, lvl_key)] = text

        else:
            # CSV / TSV
            import csv as _csv

            text = path.read_text(encoding="utf-8", errors="replace")
            sample = text[:8192]
            delimiter = ","
            if "\t" in sample and sample.count("\t") > sample.count(","):
                delimiter = "\t"

            with path.open(newline="", encoding="utf-8", errors="replace") as fh:
                reader = _csv.DictReader(fh, delimiter=delimiter)
                for r in reader:
                    dom = (
                        r.get("Parent-Friendly Domain")
                        or r.get("parent-friendly domain")
                        or r.get("Domain")
                        or r.get("domain")
                    )
                    lvl = (
                        r.get("Concern Level")
                        or r.get("concern level")
                        or r.get("Concern")
                        or r.get("concern")
                    )
                    rec = (
                        r.get("Recommendations, Supports, and Resources (Ireland – Cork/Kerry)")
                        or r.get("recommendations")
                        or r.get("Recommendations")
                    )
                    if not dom or not lvl or not rec:
                        continue
                    dom_key = _norm_domain_for_lookup(str(dom))
                    lvl_key = _norm_concern_for_lookup(str(lvl))
                    _RECS_CACHE[(dom_key, lvl_key)] = str(rec).strip()

    except Exception as e:
        current_app.logger.exception("Failed to load recommendations table: %s", e)

    return _RECS_CACHE


def _build_recommendations_from_domains(
    domain_objects: Dict[str, Dict[str, Any]],
    mode: str = "frequency",
) -> List[str]:
    """
    Build recommendations using either:
      - mode="frequency": freq_percent
      - mode="quotes":    quote_share (based on YES quotes only)

    Thresholds (same as traffic lights):
        > 70      -> "hi"   (High concern)
        >= 30     -> "some" (Some concern)
        < 30      -> "no"   (No concern, recommendation suppressed)
    """
    table = _load_recommendations_table()
    if not table:
        return []

    lines: List[str] = []
    for domain_name, dobj in domain_objects.items():
        # Choose the metric based on mode
        if mode == "quotes":
            metric = dobj.get("quote_share", 0.0)
        else:
            metric = dobj.get("freq_percent", 0.0)

        lvl_code = _percent_to_level(metric)  # "hi" | "some" | "no"

        # Do NOT generate recommendations for "no concern"
        if lvl_code == "no":
            continue

        dom_key = _norm_domain_for_lookup(domain_name)
        rec_text = table.get((dom_key, lvl_code))
        if not rec_text:
            continue

        lines.append(f"{domain_name}: {rec_text}")

    return lines


# ----------------------- Header extraction -----------------------


def _extract_header_from_transcript(session_id: str) -> Dict[str, str]:
    base = DATA_ROOT / session_id
    transcript_path = base / "transcript.txt"
    out: Dict[str, str] = {}
    if not transcript_path.exists():
        return out
    txt = transcript_path.read_text(encoding="utf-8", errors="replace")
    patterns = {
        "child_name": r"(?:child name|name of child|patient name)\s*[:\-]\s*(.+)",
        "child_age": r"\b(?:age)\s*[:\-]\s*([0-9]{1,2}\s*(?:years|yrs|y)?|[0-9]{1,2})",
        "date_of_interview": r"(?:date of interview|interview date|date)\s*[:\-]\s*([0-9]{1,2}\s*\w+\s*[0-9]{2,4}|\w+\s*[0-9]{1,2},?\s*[0-9]{4})",
        "parent_respondent": r"(?:parent respondent|respondent|parent)\s*[:\-]\s*(.+)",
        "interviewer": r"(?:interviewer)\s*[:\-]\s*(.+)",
        "referral_source": r"(?:referral source|referred by|referral)\s*[:\-]\s*(.+)",
    }
    for key, pat in patterns.items():
        m = re.search(pat, txt, flags=re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            val = re.sub(r"[\r\n]+", " ", val)
            out[key] = val
    if "date_of_interview" not in out:
        m = re.search(r"([0-9]{1,2}\s+\w+\s+[0-9]{4})", txt)
        if m:
            out["date_of_interview"] = m.group(1).strip()
    return out


def load_header_metadata(session_id: str) -> Dict[str, str]:
    meta_path = DATA_ROOT / session_id / "report_meta.json"
    header = {
        "child_name": "—",
        "child_age": "—",
        "date_of_interview": datetime.utcnow().strftime("%d %b %Y"),
        "parent_respondent": "—",
        "interviewer": "—",
        "referral_source": "—",
        "report_title": "Parent Telephone Interview Summary",
        # NEW
        "child_class_year": "—",
    }
    if meta_path.exists():
        try:
            data = json.loads(meta_path.read_text(encoding="utf-8"))
            header.update({k: v for k, v in data.items() if v is not None})
            # ensure child_class_year exists even if not in file
            header.setdefault("child_class_year", "—")
            return header
        except Exception:
            current_app.logger.warning(
                "Failed to parse report_meta.json; falling back to transcript."
            )
    extracted = _extract_header_from_transcript(session_id)
    if extracted:
        header.update({k: v for k, v in extracted.items() if v})
        header.setdefault("child_class_year", "—")
        return header
    session_dir = DATA_ROOT / session_id
    if session_dir.exists():
        try:
            ts = session_dir.stat().st_mtime
            header["date_of_interview"] = datetime.utcfromtimestamp(ts).strftime("%d %b %Y")
        except Exception:
            pass
    header.setdefault("child_class_year", "—")
    return header


# ----------------------- Shared processing helper -----------------------
def _extract_basic_info_from_rows_list(rows_list: List[Dict[str, Any]]) -> Dict[str, str]:
    """
    Scan the raw Excel rows for the basic Q&A items and pull their answers:

      - What is your child's full name?        -> child_name
      - How old is your child?                -> child_age
      - What is your full name?               -> parent_respondent
      - Who is conducting today's interview?  -> interviewer
      - How were you referred to our service? -> referral_source
      - What class/year is your child in?     -> child_class_year
    """
    info = {
        "child_name": None,
        "child_age": None,
        "parent_respondent": None,
        "interviewer": None,
        "referral_source": None,
        "child_class_year": None,
    }

    def get_any(r: Dict[str, Any], *names):
        for n in names:
            if n in r and r[n] not in (None, ""):
                return r[n]
        return None

    for r in rows_list:
        label = get_any(
            r,
            "row_label",
            "label",
            "row label",
            "rowlabel",
            "possible parent-friendly question",
            "question",
        )
        answer = get_any(
            r,
            "answer",
            "value",
            "filled_value",
            "response",
            "result",
        )
        if not label or answer in (None, ""):
            continue

        s = str(label).strip().lower().replace("’", "'")

        if "what is your child's full name" in s or "what is your childs full name" in s:
            if not info["child_name"]:
                info["child_name"] = str(answer).strip()
        elif "how old is your child" in s:
            if not info["child_age"]:
                info["child_age"] = str(answer).strip()
        elif "what is your full name" in s:
            if not info["parent_respondent"]:
                info["parent_respondent"] = str(answer).strip()
        elif "who is conducting today's interview" in s or "who is conducting today" in s:
            if not info["interviewer"]:
                info["interviewer"] = str(answer).strip()
        elif "how were you referred to our service" in s or "referred to our service" in s:
            if not info["referral_source"]:
                info["referral_source"] = str(answer).strip()
        elif "class/year" in s or "class / year" in s or "what class" in s or "what year" in s:
            if not info["child_class_year"]:
                info["child_class_year"] = str(answer).strip()

    return info


def _process_rows_into_response(
    mapped_rows: Dict[str, Dict[str, Any]],
    rows_list: List[Dict[str, Any]],
    header_meta: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """
    Build the identical 'rows_out' and 'domain_scores' payload that api_generate returns,
    given already-mapped canonical rows (mapped_rows) and the raw rows_list.

    NEW:
    - Enrich header from Excel rows where possible (child_name, parent, etc.).
    - Auto-generate a natural-language summary_overview string using Example 1 template:
        [Strengths], [Concern areas], [Class/year], [Domains of no concerns]
    """
    # ---------------- Normalise canonical rows ----------------
    normalized: Dict[str, Dict[str, Any]] = {}
    for k in CANONICAL_KEYS:
        if k in mapped_rows:
            normalized[k] = mapped_rows[k]
            continue

        found = None
        for mk in mapped_rows.keys():
            if k in mk:
                found = mk
                break
        if found:
            normalized[k] = mapped_rows[found]
            continue

        words = k.replace("_", " ").split()
        best = None
        for mk in mapped_rows.keys():
            if all(w.lower() in mk.lower() for w in words[:2]):
                best = mk
                break
        if best:
            normalized[k] = mapped_rows[best]
            continue

        normalized[k] = _empty_row()

    # ---------------- Aggregate domain metrics ----------------
    agg = _aggregate_domain_weights_from_rows(rows_list)

    # Build canonical domain objects with both frequency (percent) and quote metrics
    domain_objects: Dict[str, Dict[str, Any]] = {}
    for domain in DOMAIN_CANON:
        info = agg.get(domain, {})
        rep_val = None
        rep_quote = None

        # Find any canonical domain_* row that matches this domain
        for k, v in normalized.items():
            if not k.startswith("domain_"):
                continue
            if domain.lower().replace(" ", "_") in k:
                rep = v or {}
                rep_val = rep.get("value") or rep_val
                rep_quote = rep.get("quote") or rep_quote

        # Clean out any 'inferred' text from value / quote
        rep_val = _strip_inferred_text(rep_val)
        rep_quote = _strip_inferred_text(rep_quote)

        # Representative quote = first YES quote (if we have any)
        yes_quotes = info.get("yes_quotes", []) or []
        if not rep_quote and yes_quotes:
            rep_quote = _strip_inferred_text(yes_quotes[0])

        domain_objects[domain] = {
            "value": rep_val,
            "quote": rep_quote,
            "confidence": 1.0 if (rep_val or rep_quote) else 0.0,
            "freq_percent": info.get("percent", 0.0),
            "freq_share": info.get("freq_share", 0.0),
            "quote_count": info.get("quote_count", 0),
            "quote_yes_count": info.get("quote_yes_count", 0),
            "quote_no_count": info.get("quote_no_count", 0),
            "quote_share": info.get("quote_share", 0.0),
            # explicit lists for UI (already filtered for 'inferred')
            "yes_quotes": yes_quotes,
            "no_quotes": info.get("no_quotes", []) or [],
            # keep legacy fields but only from YES quotes
            "unique_quotes": yes_quotes,
            "all_quotes_text": info.get("quote_total_text", ""),
        }

    # ---------------- Traffic lists (frequency + quotes) ----------------
    traffic_freq_hi: List[str] = []
    traffic_freq_some: List[str] = []
    traffic_freq_no: List[str] = []

    traffic_quote_hi: List[str] = []
    traffic_quote_some: List[str] = []
    traffic_quote_no: List[str] = []

    for domain, dobj in domain_objects.items():
        p = dobj.get("freq_percent", 0.0)
        lvl = _percent_to_level(p)
        if lvl == "hi":
            traffic_freq_hi.append(domain)
        elif lvl == "some":
            traffic_freq_some.append(domain)
        else:
            traffic_freq_no.append(domain)

        qs = dobj.get("quote_share", 0.0)
        lvl_q = _percent_to_level(qs)
        if lvl_q == "hi":
            traffic_quote_hi.append(domain)
        elif lvl_q == "some":
            traffic_quote_some.append(domain)
        else:
            traffic_quote_no.append(domain)

    # ---------------- Base rows_out map ----------------
    rows_out: Dict[str, Any] = {}
    for k in CANONICAL_KEYS:
        rows_out[k] = normalized.get(k, _empty_row())

    # Insert detailed domain_* objects (with shares and quotes) for key domains
    for domain_key, domain_name in {
        "domain_attention_adhd": "Attention / ADHD",
        "domain_learning_dyslexia": "Learning / Dyslexia",
        "domain_sleep": "Sleep",
        "domain_motor_skills": "Motor Skills",
        "domain_anxiety_emotion": "Anxiety / Emotion",
        "domain_social_communication": "Social / Communication",
        "domain_eating": "Eating",
    }.items():
        dobj = domain_objects.get(domain_name, {})
        rows_out[domain_key] = {
            "value": dobj.get("value"),
            "quote": dobj.get("quote"),
            "confidence": dobj.get("confidence", 0.0),
            "freq_percent": dobj.get("freq_percent", 0.0),
            "freq_share": dobj.get("freq_share", 0.0),
            "quote_count": dobj.get("quote_count", 0),
            "quote_yes_count": dobj.get("quote_yes_count", 0),
            "quote_no_count": dobj.get("quote_no_count", 0),
            "quote_share": dobj.get("quote_share", 0.0),
            "yes_quotes": dobj.get("yes_quotes", []),
            "no_quotes": dobj.get("no_quotes", []),
            "unique_quotes": dobj.get("unique_quotes", []),
            "all_quotes_text": dobj.get("all_quotes_text", ""),
        }

    # Helper to wrap traffic lists
    def _to_wrap_list(arr: List[str]) -> Dict[str, Any]:
        if not arr:
            return {"value": None, "quote": None, "confidence": 0.0, "list": [], "text": None}
        return {
            "value": ", ".join(arr),
            "quote": None,
            "confidence": 1.0,
            "list": arr,
            "text": ", ".join(arr),
        }

    # Frequency traffic
    rows_out["traffic_freq_high"] = _to_wrap_list(traffic_freq_hi)
    rows_out["traffic_freq_some"] = _to_wrap_list(traffic_freq_some)
    rows_out["traffic_freq_no"] = _to_wrap_list(traffic_freq_no)

    # Quote traffic
    rows_out["traffic_quote_high"] = _to_wrap_list(traffic_quote_hi)
    rows_out["traffic_quote_some"] = _to_wrap_list(traffic_quote_some)
    rows_out["traffic_quote_no"] = _to_wrap_list(traffic_quote_no)

    # Legacy aliases
    rows_out["traffic_high"] = rows_out["traffic_freq_high"]
    rows_out["traffic_some"] = rows_out["traffic_freq_some"]
    rows_out["traffic_no"] = rows_out["traffic_freq_no"]

    # Domain scores for charting
    domain_scores: Dict[str, Any] = {}
    for domain in DOMAIN_CANON:
        info = agg.get(domain, {})
        domain_scores[domain] = {
            "freq_percent": info.get("percent", 0.0),
            "freq_share": info.get("freq_share", 0.0),
            "quote_share": info.get("quote_share", 0.0),
            "quote_count": info.get("quote_count", 0),
            "quote_yes_count": info.get("quote_yes_count", 0),
            "quote_no_count": info.get("quote_no_count", 0),
        }

    rows_out["domain_objects"] = domain_objects
    rows_out["domain_scores"] = domain_scores

    # ---------------- Header: meta first, then gentle Q&A backfill ----------------
    # 1) Start from header_meta if provided (report_meta.json / upload header)
    if header_meta:
        header = dict(header_meta)
    else:
        header = {
            "child_name": "—",
            "child_age": "—",
            "date_of_interview": datetime.utcnow().strftime("%d %b %Y"),
            "parent_respondent": "—",
            "interviewer": "—",
            "referral_source": "—",
            "report_title": "Parent Telephone Interview Summary",
            "child_class_year": "—",
        }

    # Ensure required keys exist
    header.setdefault("child_name", "—")
    header.setdefault("child_age", "—")
    header.setdefault(
        "date_of_interview",
        datetime.utcnow().strftime("%d %b %Y"),
    )
    header.setdefault("parent_respondent", "—")
    header.setdefault("interviewer", "—")
    header.setdefault("referral_source", "—")
    header.setdefault("report_title", "Parent Telephone Interview Summary")
    header.setdefault("child_class_year", "—")

    # 2) NEW: backfill ONLY from the structured Q&A rows
    try:
        basic = _extract_basic_info_from_rows_list(rows_list)
        for hk, val in basic.items():
            if not val:
                continue
            cleaned = _strip_inferred_text(val)
            if cleaned and header.get(hk) in (None, "", "—"):
                header[hk] = cleaned
    except Exception as e:
        current_app.logger.warning("Header backfill from rows_list failed: %s", e)

    # ---------------- Build natural-language summary (Example 1 template) ----------------
    def _get_domain_text(domain_name: str) -> str:
        # 1. Try domain_objects (quote > value)
        dobj = domain_objects.get(domain_name, {}) or {}
        txt = dobj.get("quote") or dobj.get("value")

        # 2. If still empty, look up canonical domain_* row in rows_out
        if not txt:
            domain_key_map = {
                "Parent's Goal for this Assessment": "domain_parents_goal",
                "Strengths / Interests": "domain_strengths_interests",
            }
            dk = domain_key_map.get(domain_name)
            if dk and dk in rows_out:
                row = rows_out[dk] or {}
                txt = row.get("quote") or row.get("value")

        cleaned = _strip_inferred_text(txt)
        return cleaned or ""

    child_name = header.get("child_name") or "your child"
    child_age = header.get("child_age") or "of school age"
    class_year = header.get("child_class_year") or "their current class"

    # [Strengths]
    strengths_text = _get_domain_text("Strengths / Interests")

    # [Concern areas] – domains raised (high / some) using frequency traffic
    concern_domains = traffic_freq_hi + traffic_freq_some
    if concern_domains:
        concern_areas = ", ".join(concern_domains)
    else:
        concern_areas = "the areas discussed"

    # [Domains of no concerns] – domains with 'no concern'
    if traffic_freq_no:
        domains_no_concern = ", ".join(traffic_freq_no)
    else:
        domains_no_concern = "no specific domains"

    if strengths_text:
        strengths_clause = (
            f"You described many lovely strengths, including {strengths_text}, and it’s clear "
            f"how bright and engaging {child_name} is."
        )
    else:
        strengths_clause = (
            f"You described many lovely strengths, and it’s clear how bright and engaging {child_name} is."
        )

    summary_sentences = [
        f"Thank you for sharing your thoughts about {child_name}. {child_name} is {child_age} and lives at home with you and your family.",
        strengths_clause,
        f"You also mentioned some worries in the areas of {concern_areas}, and understanding these helps us think about how best to support {child_name}.",
        f"It’s very positive that you feel well connected as a family and that {child_name} is surrounded by such strong support.",
        f"{child_name} is currently in {class_year} at school, and you noted no concerns in {domains_no_concern}, which is reassuring.",
        "The recommendations below may offer helpful ideas for the next steps.",
    ]

    summary_text = " ".join(s.strip() for s in summary_sentences if s and s.strip())

    rows_out["summary_overview"] = {
        "value": summary_text if summary_text else None,
        "quote": None,
        "confidence": 1.0 if summary_text else 0.0,
    }

    # ---------------- Recommendations from domain scores ----------------
    rec_lines_freq = _build_recommendations_from_domains(
        domain_objects, mode="frequency"
    )
    rec_lines_quotes = _build_recommendations_from_domains(
        domain_objects, mode="quotes"
    )

    if rec_lines_freq:
        rows_out["recommendations_freq"] = {
            "value": "\n".join(rec_lines_freq),
            "quote": None,
            "confidence": 1.0,
            "list": rec_lines_freq,
        }

    if rec_lines_quotes:
        rows_out["recommendations_quotes"] = {
            "value": "\n".join(rec_lines_quotes),
            "quote": None,
            "confidence": 1.0,
            "list": rec_lines_quotes,
        }

    # default "recommendations" = frequency list, then quotes, else empty
    if rec_lines_freq:
        rows_out["recommendations"] = rows_out["recommendations_freq"]
    elif rec_lines_quotes:
        rows_out["recommendations"] = rows_out["recommendations_quotes"]
    elif "recommendations" not in rows_out or not rows_out["recommendations"].get("value"):
        rows_out["recommendations"] = {
            "value": None,
            "quote": None,
            "confidence": 0.0,
        }

    # ---------------- Persist derived rows (optional) ----------------
    try:
        uid = str(uuid.uuid4())[:8]
        outp = DATA_ROOT / f"upload_{uid}_analytics_rows.json"
        outp.write_text(json.dumps(rows_out, ensure_ascii=False, indent=2), encoding="utf-8")
        current_app.logger.info("Analytics: wrote derived uploaded rows to %s", outp)
    except Exception as e:
        current_app.logger.warning("Analytics: could not persist derived rows for upload: %s", e)

    return {"ok": True, "header": header, "rows": rows_out}


def _rewrite_summary_with_llm(
    summary_text: str,
    header: Optional[Dict[str, str]] = None,
    domain_objects: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Rewrite the summary in a clinician-written style using Gemini.
    Falls back to the original text if API key missing or any error occurs.
    """
    summary_text = (summary_text or "").strip()
    if not summary_text:
        return summary_text

    # Choose a working Gemini model dynamically
    model_name = _get_gemini_model_name()
    if not model_name:
        # Already logged; just fall back
        return summary_text

    try:
        model = GenerativeModel(model_name)
    except Exception as e:
        current_app.logger.exception("Gemini: failed to initialise GenerativeModel(%s): %s", model_name, e)
        return summary_text

    header = header or {}
    child_name = header.get("child_name") or "the child"
    child_age = header.get("child_age") or "—"
    parent_name = header.get("parent_respondent") or "the parent"
    interviewer = header.get("interviewer") or "the clinician"

    # Optional domain context
    top_domains = []
    if domain_objects:
        try:
            scored = sorted(
                domain_objects.items(),
                key=lambda kv: float(kv[1].get("freq_percent", 0.0) or 0.0),
                reverse=True,
            )
            for name, dobj in scored[:4]:
                top_domains.append(f"- {name}: {dobj.get('quote') or dobj.get('value') or ''}")
        except Exception:
            pass

    # Build context
    context_lines = [
        f"Child: {child_name}, age: {child_age}.",
        f"Parent respondent: {parent_name}.",
        f"Interviewer: {interviewer}.",
    ]
    if top_domains:
        context_lines.append("Key domains mentioned:")
        context_lines.extend(top_domains)

    context_block = "\n".join(context_lines)

    system_prompt = (
        "You are a paediatric clinician writing a brief, clear, parent-facing summary. "
        "Use UK/Irish English, warm professional tone, and avoid jargon. "
        "Do NOT introduce any new diagnoses or assumptions. "
        "Stay within 150–220 words. Keep content accurate and empathetic."
    )

    user_prompt = (
        f"{system_prompt}\n\n"
        "Rewrite the following summary into a polished, clinician-written version.\n\n"
        "=== CONTEXT ===\n"
        f"{context_block}\n\n"
        "=== ORIGINAL SUMMARY ===\n"
        f"{summary_text}\n\n"
        "=== TASK ===\n"
        "Rewrite as a single coherent paragraph (or two short paragraphs). "
        "Do not add new information, diagnoses, or headings."
    )

    try:
        response = model.generate_content(user_prompt)
        new_text = (response.text or "").strip()
        if not new_text:
            current_app.logger.warning("Gemini rewrite returned empty text; falling back to original.")
            return summary_text
        return new_text
    except Exception as e:
        current_app.logger.exception("Gemini rewrite failed: %s", e)
        return summary_text


# ----------------------- NEW: LLM helpers for disruption flags -----------------------


def _llm_annotate_disruptions(item_text: str) -> Dict[str, str]:
    """
    Given the free-text description of a problem/behaviour (e.g. an item quote),
    return y/n flags for functional impact dimensions:

        - disrupts_eating_feeding
        - disrupts_social_activity
        - disrupts_major_role_obligations
        - disrupts_education
        - disrupts_social_obligations
        - disrupts_sleep
        - disrupts_physical_activity
        - disrupts_other_everyday_functioning

    Uses Gemini. If the API key is missing or anything fails, returns all 'n'.
    """
    # default: assume no disruption
    default_flags = {
        "disrupts_eating_feeding": "n",
        "disrupts_social_activity": "n",
        "disrupts_major_role_obligations": "n",
        "disrupts_education": "n",
        "disrupts_social_obligations": "n",
        "disrupts_sleep": "n",
        "disrupts_physical_activity": "n",
        "disrupts_other_everyday_functioning": "n",
    }

    item_text = (item_text or "").strip()
    if not item_text:
        return default_flags

    # 🔹 Use dynamic model selection (from the helper we added earlier)
    model_name = _get_gemini_model_name()
    if not model_name:
        return default_flags

    try:
        model = GenerativeModel(model_name)
    except Exception as e:
        current_app.logger.exception("Gemini: failed to initialise GenerativeModel(%s): %s", model_name, e)
        return default_flags

    # 🔹 LLM input EXACTLY as you specified
    user_prompt = (
        "Every item can have:\n"
        "- Disrupts eating/feeding (y/n)\n"
        "- Disrupts social activity (y/n)\n"
        "- Disrupts fulfillment of major role obligations "
        "(forgetting homework, being respectful to parents and teachers, "
        "preparing food for siblings, etc.) (y/n)\n"
        "- Disrupts education (y/n)\n"
        "- Disrupts fulfillment of social obligations (y/n)\n"
        "- Disrupts sleep (y/n)\n"
        "- Disrupts physical activity (y/n)\n"
        "- Disrupts other everyday functioning "
        "(using public transport, toileting, hygiene etc.) (y/n)\n\n"
        "Given the behaviour / difficulty below, decide if it clearly disrupts each area.\n"
        "If the text is unclear or the area is not mentioned, answer 'n' for that area.\n\n"
        "Text to analyse:\n"
        f"\"\"\"{item_text}\"\"\"\n\n"
        "Return STRICTLY valid JSON in this shape (no extra text):\n"
        "{\n"
        '  \"disrupts_eating_feeding\": \"y|n\",\n'
        '  \"disrupts_social_activity\": \"y|n\",\n'
        '  \"disrupts_major_role_obligations\": \"y|n\",\n'
        '  \"disrupts_education\": \"y|n\",\n'
        '  \"disrupts_social_obligations\": \"y|n\",\n'
        '  \"disrupts_sleep\": \"y|n\",\n'
        '  \"disrupts_physical_activity\": \"y|n\",\n'
        '  \"disrupts_other_everyday_functioning\": \"y|n\"\n'
        "}\n"
    )

    try:
        resp = model.generate_content(user_prompt)
        raw = (resp.text or "").strip()

        try:
            parsed = json.loads(raw)
        except Exception:
            # Sometimes models wrap JSON in markdown fences – try to strip them
            m = re.search(r"\{.*\}", raw, flags=re.DOTALL)
            if not m:
                current_app.logger.warning("Failed to parse LLM disruption JSON: %s", raw[:200])
                return default_flags
            parsed = json.loads(m.group(0))

        out = dict(default_flags)
        for k in default_flags.keys():
            val = str(parsed.get(k, "n")).strip().lower()
            out[k] = "y" if val in ("y", "yes", "true", "1") else "n"

        return out

    except Exception as e:
        current_app.logger.exception("Gemini disruption annotation failed: %s", e)
        return default_flags

def annotate_row_with_disruptions_for_excel(row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convenience helper for the Excel/CSV export pipeline.

    - Takes a row dict (e.g. with 'value' and/or 'quote').
    - Builds a text snippet.
    - Calls the LLM to classify functional disruption.
    - Writes BOTH:
        * internal keys (disrupts_...),
        * and human-friendly CSV/Excel column headers:

            - eating/feeding (y/n)
            - social activity (y/n)
            - fulfillment of major role obligations (y/n)
            - education (y/n)
            - fulfillment of social obligations (y/n)
            - sleep (y/n)
            - physical activity (y/n)
            - other everyday functioning (y/n)

      into the row dict, so csv.DictWriter / Excel export can just dump the row.
    """
    if not isinstance(row, dict):
        row = {}

    # Use quote first (richer description), then value/question
    text_pieces = []
    if row.get("quote"):
        text_pieces.append(str(row["quote"]))
    if row.get("value"):
        text_pieces.append(str(row["value"]))
    # Optional: include any 'question' or 'label' to give more context
    for k in ("row_label", "label", "question", "possible parent-friendly question"):
        if k in row and row[k]:
            text_pieces.append(str(row[k]))

    combined_text = " ".join(text_pieces).strip()
    flags = _llm_annotate_disruptions(combined_text)

    # Attach canonical internal keys (if you want to use them programmatically)
    for k, v in flags.items():
        row[k] = v

    # Attach HUMAN-FRIENDLY COLUMN NAMES for CSV/Excel
    header_map = {
        "eating/feeding (y/n)": "disrupts_eating_feeding",
        "social activity (y/n)": "disrupts_social_activity",
        "fulfillment of major role obligations (y/n)": "disrupts_major_role_obligations",
        "education (y/n)": "disrupts_education",
        "fulfillment of social obligations (y/n)": "disrupts_social_obligations",
        "sleep (y/n)": "disrupts_sleep",
        "physical activity (y/n)": "disrupts_physical_activity",
        "other everyday functioning (y/n)": "disrupts_other_everyday_functioning",
    }

    for col_name, flag_key in header_map.items():
        row[col_name] = flags.get(flag_key, "n")

    return flags

@analytics_bp.post("/api/analytics/<session_id>/summary/rewrite")
def api_rewrite_summary(session_id):
    """
    Rewrites the current 'Summary of Key Findings' using the LLM, as if written by a clinician.
    Expects JSON body with at least { "summary": "..." }.
    Optionally also accepts header fields (child_name, child_age, etc.).
    """
    data = request.get_json(silent=True) or {}
    summary = (data.get("summary") or "").strip()
    if not summary:
        return jsonify({"ok": False, "error": "Missing summary text"}), 400

    header = {
        "child_name": data.get("child_name"),
        "child_age": data.get("child_age"),
        "parent_respondent": data.get("parent_respondent"),
        "interviewer": data.get("interviewer"),
    }

    # We *could* also pass domain_objects from disk, but the simple form is:
    new_summary = _rewrite_summary_with_llm(summary, header=header, domain_objects=None)

    return jsonify({"ok": True, "summary": new_summary})


# ----------------------- Upload endpoint -----------------------


@analytics_bp.post("/api/analytics/upload")
def api_upload_and_generate():
    """
    Accepts a multipart-form file upload with key 'file'.
    Supports csv/tsv/txt/json/xlsx.
    Returns identical payload shape to /api/analytics/<session_id>.
    Uploaded file is removed after processing (best-effort).
    """
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded (use key 'file')"}), 400
    f = request.files["file"]
    if f.filename == "":
        return jsonify({"ok": False, "error": "Empty filename"}), 400

    filename = f.filename
    suffix = Path(filename).suffix.lower() or ".csv"
    tmp_path: Optional[Path] = None
    try:
        # write to temp file to reuse existing readers
        uid = uuid.uuid4().hex
        tmp_dir = DATA_ROOT / "uploads"
        tmp_dir.mkdir(parents=True, exist_ok=True)
        tmp_path = tmp_dir / f"{uid}{suffix}"
        f.save(str(tmp_path))

        # read rows_list using appropriate reader
        if suffix in (".csv", ".txt", ".tsv"):
            rows_list = _read_csv_filled(tmp_path)
        elif suffix in (".json",):
            rows_list = _read_json_filled(tmp_path)
        elif suffix in (".xlsx", ".xls"):
            rows_list = _read_xlsx_filled(tmp_path)
        else:
            rows_list = _read_csv_filled(tmp_path)
    except Exception as e:
        current_app.logger.exception("Failed to save/read uploaded file: %s", e)
        try:
            if tmp_path and tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass
        return jsonify({"ok": False, "error": "Failed to save or parse uploaded file"}), 500

    mapped = _map_row_list_to_dict(rows_list)
    header_meta = {
        "child_name": "—",
        "child_age": "—",
        "date_of_interview": datetime.utcnow().strftime("%d %b %Y"),
        "parent_respondent": "Uploaded CSV",
        "interviewer": "—",
        "referral_source": Path(filename).name,
        "report_title": "Uploaded Data Analytics",
        "child_class_year": "—",
    }

    try:
        response_payload = _process_rows_into_response(mapped, rows_list, header_meta)
        return jsonify(response_payload)
    finally:
        try:
            if tmp_path and tmp_path.exists():
                tmp_path.unlink()
                current_app.logger.info("Analytics: removed uploaded temp file %s", tmp_path)
        except Exception as e:
            current_app.logger.warning(
                "Analytics: failed to remove uploaded temp file %s: %s", tmp_path, e
            )


# ----------------------- API (pure-Python) -----------------------


@analytics_bp.get("/api/analytics/<session_id>")
def api_generate(session_id):
    rows, path_used, raw_rows_list = load_filled_rows_for_session(session_id)
    if not rows:
        current_app.logger.error(
            "Analytics: no filled rows found for session %s (path=%s)", session_id, path_used
        )
        return jsonify({"ok": False, "error": "Missing filled file or no recognizable rows"}), 400

    response_payload = _process_rows_into_response(
        rows, raw_rows_list, load_header_metadata(session_id)
    )
    return jsonify(response_payload)


# ---- Page ----
@analytics_bp.get("/analytics/<session_id>")
def page_analytics(session_id):
    header = load_header_metadata(session_id)
    return render_template("analytics.html", session_id=session_id, header=header)
